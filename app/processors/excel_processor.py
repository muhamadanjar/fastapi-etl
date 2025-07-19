# ==============================================
# app/processors/excel_processor.py
# ==============================================
import pandas as pd
import openpyxl
import xlrd
from typing import Dict, List, Any, Optional, Tuple, Iterator
from pathlib import Path
from datetime import datetime
import json

from .base_processor import BaseProcessor
from app.core.exceptions import FileProcessingException
from app.infrastructure.db.models.raw_data.file_registry import FileRegistry
from app.utils.logger import get_logger

logger = get_logger(__name__)

class ExcelProcessor(BaseProcessor):
    """
    Excel file processor supporting both XLSX and XLS formats
    with multi-sheet processing, formula handling, and data type detection
    """
    
    def __init__(self, db_session, batch_id: Optional[str] = None, **kwargs):
        """
        Initialize Excel processor
        
        Args:
            db_session: Database session
            batch_id: Optional batch ID
            **kwargs: Additional configuration options
        """
        super().__init__(db_session, batch_id)
        
        # Excel-specific configuration
        self.sheet_names = kwargs.get('sheet_names', None)  # Process all sheets if None
        self.header_row = kwargs.get('header_row', 0)      # First row as header by default
        self.skip_rows = kwargs.get('skip_rows', None)     # Rows to skip from top
        self.max_rows = kwargs.get('max_rows', None)       # Maximum rows to read
        self.chunk_size = kwargs.get('chunk_size', 10000)  # Chunk size for large files
        self.ignore_hidden_sheets = kwargs.get('ignore_hidden_sheets', True)
        self.preserve_formulas = kwargs.get('preserve_formulas', False)
        self.include_formatting = kwargs.get('include_formatting', False)
        
    async def validate_file_format(self, file_path: str) -> Tuple[bool, str]:
        """
        Validate if file is a valid Excel format
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        try:
            file_path_obj = Path(file_path)
            
            # Check file exists
            if not file_path_obj.exists():
                return False, "File does not exist"
            
            # Check file extension
            valid_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']
            if file_path_obj.suffix.lower() not in valid_extensions:
                return False, f"Invalid file extension: {file_path_obj.suffix}. Expected: {valid_extensions}"
            
            # Check file size
            file_size = file_path_obj.stat().st_size
            if file_size == 0:
                return False, "File is empty"
            
            if file_size > 500 * 1024 * 1024:  # 500MB limit for Excel
                return False, "File too large (>500MB)"
            
            # Try to open the Excel file
            try:
                if file_path_obj.suffix.lower() == '.xls':
                    # Handle legacy XLS format
                    workbook = xlrd.open_workbook(file_path)
                    if workbook.nsheets == 0:
                        return False, "No sheets found in Excel file"
                else:
                    # Handle modern XLSX format
                    workbook = openpyxl.load_workbook(file_path, read_only=True)
                    if len(workbook.sheetnames) == 0:
                        return False, "No sheets found in Excel file"
                
                return True, "Valid Excel file"
                
            except Exception as e:
                return False, f"Cannot open Excel file: {str(e)}"
            
        except Exception as e:
            return False, f"Excel validation error: {str(e)}"
    
    async def detect_structure(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Detect Excel structure including sheets, columns, and data types
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet and column information
        """
        try:
            structure_info = []
            
            # Get sheet information
            sheets_info = await self._get_sheets_info(file_path)
            
            for sheet_info in sheets_info:
                sheet_name = sheet_info['name']
                
                try:
                    # Read sample data from sheet
                    df_sample = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=self.header_row,
                        nrows=1000,  # Sample first 1000 rows
                        skiprows=self.skip_rows
                    )
                    
                    # Skip empty sheets
                    if df_sample.empty:
                        self.logger.warning(f"Sheet '{sheet_name}' is empty, skipping")
                        continue
                    
                    columns_info = []
                    
                    for position, column_name in enumerate(df_sample.columns):
                        column_data = df_sample[column_name]
                        
                        # Handle unnamed columns
                        if pd.isna(column_name) or str(column_name).startswith('Unnamed'):
                            column_name = f"Column_{position + 1}"
                        
                        # Get sample values (non-null)
                        sample_values = column_data.dropna().head(10).astype(str).tolist()
                        
                        # Calculate statistics
                        null_count = column_data.isnull().sum()
                        unique_count = column_data.nunique()
                        
                        # Detect data type
                        data_type = self._detect_data_type(sample_values)
                        
                        # Calculate string lengths for string columns
                        min_length = None
                        max_length = None
                        if data_type == "STRING":
                            str_lengths = [len(str(val)) for val in sample_values if val]
                            if str_lengths:
                                min_length = min(str_lengths)
                                max_length = max(str_lengths)
                        
                        column_info = {
                            "name": str(column_name),
                            "position": position,
                            "data_type": data_type,
                            "sample_values": sample_values,
                            "null_count": int(null_count),
                            "unique_count": int(unique_count),
                            "total_count": len(column_data),
                            "min_length": min_length,
                            "max_length": max_length,
                            "null_percentage": round((null_count / len(column_data)) * 100, 2),
                            "sheet_name": sheet_name
                        }
                        
                        columns_info.append(column_info)
                    
                    # Add sheet structure info
                    sheet_structure = {
                        "sheet_name": sheet_name,
                        "columns": columns_info,
                        "total_rows": len(df_sample),
                        "total_columns": len(columns_info),
                        "is_hidden": sheet_info.get('hidden', False),
                        "sheet_type": sheet_info.get('type', 'worksheet')
                    }
                    
                    structure_info.append(sheet_structure)
                    
                except Exception as e:
                    self.logger.error(f"Error analyzing sheet '{sheet_name}': {str(e)}")
                    continue
            
            self.logger.info(f"Detected structure for {len(structure_info)} sheets")
            return structure_info
            
        except Exception as e:
            self.logger.error(f"Error detecting Excel structure: {str(e)}")
            raise FileProcessingException(f"Failed to detect Excel structure: {str(e)}")
    
    async def preview_data(self, file_path: str, rows: int = 10) -> Dict[str, Any]:
        """
        Generate preview of Excel data from all sheets
        
        Args:
            file_path: Path to Excel file
            rows: Number of rows to preview per sheet
            
        Returns:
            Dictionary containing preview data and metadata
        """
        try:
            sheets_info = await self._get_sheets_info(file_path)
            preview_data = {
                "sheets": [],
                "metadata": {
                    "total_sheets": len(sheets_info),
                    "file_size": self._get_file_size(file_path),
                    "file_type": Path(file_path).suffix.upper()
                }
            }
            
            for sheet_info in sheets_info:
                sheet_name = sheet_info['name']
                
                try:
                    # Read preview data from sheet
                    df_preview = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=self.header_row,
                        nrows=rows,
                        skiprows=self.skip_rows
                    )
                    
                    if df_preview.empty:
                        continue
                    
                    # Convert to records for JSON serialization
                    preview_records = df_preview.fillna("").to_dict('records')
                    
                    # Get total row count for the sheet
                    df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    total_rows = len(df_full) - (self.header_row + 1 if self.header_row is not None else 0)
                    
                    sheet_preview = {
                        "sheet_name": sheet_name,
                        "columns": df_preview.columns.tolist(),
                        "data": preview_records,
                        "metadata": {
                            "total_rows": total_rows,
                            "preview_rows": len(preview_records),
                            "total_columns": len(df_preview.columns),
                            "is_hidden": sheet_info.get('hidden', False),
                            "sheet_type": sheet_info.get('type', 'worksheet')
                        }
                    }
                    
                    preview_data["sheets"].append(sheet_preview)
                    
                except Exception as e:
                    self.logger.error(f"Error generating preview for sheet '{sheet_name}': {str(e)}")
                    continue
            
            return preview_data
            
        except Exception as e:
            self.logger.error(f"Error generating Excel preview: {str(e)}")
            raise FileProcessingException(f"Failed to generate Excel preview: {str(e)}")
    
    async def process_file(self, file_path: str, file_registry: FileRegistry) -> Dict[str, Any]:
        """
        Process Excel file and extract data from all sheets
        
        Args:
            file_path: Path to Excel file
            file_registry: File registry record
            
        Returns:
            Processing results and statistics
        """
        try:
            self.logger.info(f"Starting Excel processing for file: {file_path}")
            
            # Get sheet information
            sheets_info = await self._get_sheets_info(file_path)
            
            # Update file metadata
            metadata = file_registry.metadata or {}
            metadata.update({
                "total_sheets": len(sheets_info),
                "sheets": [s['name'] for s in sheets_info],
                "file_type": Path(file_path).suffix.upper(),
                "processor": "ExcelProcessor"
            })
            file_registry.metadata = metadata
            self.db_session.add(file_registry)
            
            # Detect and save structure for all sheets
            structure_info = await self.detect_structure(file_path)
            
            # Flatten column info for saving
            all_columns = []
            for sheet_structure in structure_info:
                all_columns.extend(sheet_structure['columns'])
            
            await self.save_column_structure(file_registry, all_columns)
            
            # Process each sheet
            total_stats = {
                "total_records": 0,
                "successful_records": 0,
                "failed_records": 0,
                "validation_errors": [],
                "processing_time": 0,
                "sheets_processed": 0,
                "file_type": "Excel",
                "sheets_info": []
            }
            
            for sheet_info in sheets_info:
                sheet_name = sheet_info['name']
                
                # Skip hidden sheets if configured
                if self.ignore_hidden_sheets and sheet_info.get('hidden', False):
                    self.logger.info(f"Skipping hidden sheet: {sheet_name}")
                    continue
                
                try:
                    self.logger.info(f"Processing sheet: {sheet_name}")
                    
                    # Process sheet records
                    record_iterator = self._read_excel_sheet_chunks(file_path, sheet_name)
                    sheet_stats = await self.process_records(record_iterator, file_registry)
                    
                    # Update total statistics
                    total_stats["total_records"] += sheet_stats["total_records"]
                    total_stats["successful_records"] += sheet_stats["successful_records"]
                    total_stats["failed_records"] += sheet_stats["failed_records"]
                    total_stats["validation_errors"].extend(sheet_stats["validation_errors"])
                    total_stats["processing_time"] += sheet_stats["processing_time"]
                    total_stats["sheets_processed"] += 1
                    
                    # Add sheet-specific info
                    sheet_info_detail = {
                        "sheet_name": sheet_name,
                        "records": sheet_stats["total_records"],
                        "success_rate": round((sheet_stats["successful_records"] / max(sheet_stats["total_records"], 1)) * 100, 2)
                    }
                    total_stats["sheets_info"].append(sheet_info_detail)
                    
                except Exception as e:
                    self.logger.error(f"Error processing sheet '{sheet_name}': {str(e)}")
                    total_stats["validation_errors"].append(f"Sheet '{sheet_name}' processing failed: {str(e)}")
            
            self.logger.info(f"Excel processing completed: {total_stats}")
            return total_stats
            
        except Exception as e:
            self.logger.error(f"Error processing Excel file: {str(e)}")
            raise FileProcessingException(f"Failed to process Excel file: {str(e)}")
    
    async def _get_sheets_info(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Get information about all sheets in the Excel file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet information dictionaries
        """
        try:
            sheets_info = []
            file_extension = Path(file_path).suffix.lower()
            
            if file_extension == '.xls':
                # Handle legacy XLS format
                workbook = xlrd.open_workbook(file_path)
                for i, sheet_name in enumerate(workbook.sheet_names()):
                    sheet = workbook.sheet_by_index(i)
                    sheets_info.append({
                        "name": sheet_name,
                        "index": i,
                        "rows": sheet.nrows,
                        "columns": sheet.ncols,
                        "hidden": False,  # XLS doesn't easily expose hidden state
                        "type": "worksheet"
                    })
            else:
                # Handle modern XLSX format
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                for i, sheet_name in enumerate(workbook.sheetnames):
                    worksheet = workbook[sheet_name]
                    sheets_info.append({
                        "name": sheet_name,
                        "index": i,
                        "rows": worksheet.max_row,
                        "columns": worksheet.max_column,
                        "hidden": worksheet.sheet_state == 'hidden',
                        "type": "worksheet"  # Could extend to detect chart sheets, etc.
                    })
                
                workbook.close()
            
            return sheets_info
            
        except Exception as e:
            self.logger.error(f"Error getting sheets info: {str(e)}")
            raise FileProcessingException(f"Failed to get Excel sheets information: {str(e)}")
    
    def _read_excel_sheet_chunks(self, file_path: str, sheet_name: str) -> Iterator[Dict[str, Any]]:
        """
        Read Excel sheet in chunks to handle large files efficiently
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to read
            
        Yields:
            Dictionary records from Excel sheet
        """
        try:
            # For Excel, we'll read in chunks using skiprows and nrows
            total_rows = self._get_sheet_row_count(file_path, sheet_name)
            
            for start_row in range(0, total_rows, self.chunk_size):
                try:
                    # Calculate skiprows (account for header)
                    skip_rows = None
                    if start_row > 0:
                        # Skip header for subsequent chunks
                        skip_rows = list(range(self.header_row + 1)) if self.header_row is not None else None
                        actual_skip = start_row + (self.header_row + 1 if self.header_row is not None else 0)
                    else:
                        skip_rows = self.skip_rows
                        actual_skip = start_row
                    
                    # Read chunk
                    df_chunk = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=self.header_row if start_row == 0 else 0,
                        skiprows=range(1, actual_skip + 1) if actual_skip > 0 and start_row > 0 else skip_rows,
                        nrows=self.chunk_size
                    )
                    
                    if df_chunk.empty:
                        break
                    
                    # Convert chunk to records
                    chunk_records = df_chunk.fillna("").to_dict('records')
                    
                    for record in chunk_records:
                        # Clean up the record and add sheet information
                        cleaned_record = {
                            key.strip() if isinstance(key, str) else str(key): 
                            str(value).strip() if value and not pd.isna(value) else None
                            for key, value in record.items()
                        }
                        
                        # Add sheet metadata to record
                        cleaned_record['_sheet_name'] = sheet_name
                        
                        yield cleaned_record
                        
                except Exception as e:
                    self.logger.error(f"Error reading chunk starting at row {start_row}: {str(e)}")
                    continue
                    
        except Exception as e:
            self.logger.error(f"Error reading Excel sheet chunks: {str(e)}")
            raise FileProcessingException(f"Failed to read Excel sheet: {str(e)}")
    
    def _get_sheet_row_count(self, file_path: str, sheet_name: str) -> int:
        """
        Get total row count for a specific sheet
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet
            
        Returns:
            Total number of rows
        """
        try:
            file_extension = Path(file_path).suffix.lower()
            
            if file_extension == '.xls':
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_name(sheet_name)
                return sheet.nrows
            else:
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                worksheet = workbook[sheet_name]
                row_count = worksheet.max_row
                workbook.close()
                return row_count
                
        except Exception as e:
            self.logger.warning(f"Could not get row count for sheet '{sheet_name}': {str(e)}")
            return 1000000  # Large default value
    
    async def _custom_record_validation(self, record: Dict[str, Any], row_number: int) -> Dict[str, Any]:
        """
        Excel-specific record validation
        
        Args:
            record: Record to validate
            row_number: Row number for error reporting
            
        Returns:
            Validation result
        """
        validation_result = {"is_valid": True, "errors": []}
        
        try:
            # Remove sheet metadata for validation
            record_copy = {k: v for k, v in record.items() if not k.startswith('_')}
            
            # Check for records with only empty strings or whitespace
            non_empty_values = [v for v in record_copy.values() if v and str(v).strip()]
            if not non_empty_values:
                validation_result["is_valid"] = False
                validation_result["errors"].append("Record contains only empty values")
            
            # Check for Excel error values
            excel_errors = ['#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#REF!', '#VALUE!']
            for field_name, value in record_copy.items():
                if value and str(value) in excel_errors:
                    validation_result["errors"].append(f"Field '{field_name}' contains Excel error: {value}")
            
            # Check for extremely large numeric values (potential Excel precision issues)
            for field_name, value in record_copy.items():
                if value:
                    try:
                        if isinstance(value, (int, float)) and abs(value) > 1e15:
                            validation_result["errors"].append(f"Field '{field_name}' has extremely large value: {value}")
                    except (ValueError, TypeError):
                        pass
            
        except Exception as e:
            validation_result["is_valid"] = False
            validation_result["errors"].append(f"Validation error: {str(e)}")
        
        return validation_result
    