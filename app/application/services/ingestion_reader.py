from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping

import requests
from openpyxl import load_workbook

from app.application.services.dataset_engine import bounded_records
from app.schemas.workspace import ApiSourceConfig, CredentialWrite


class IngestionReadError(ValueError):
    code = "INGESTION_READ_ERROR"


def _normalize_record(value: Any) -> Dict[str, Any]:
    if not isinstance(value, Mapping):
        raise IngestionReadError("Every source record must be a JSON object")
    return {str(key): item for key, item in value.items()}


def read_file_records(path: str) -> tuple[List[Dict[str, Any]], int]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            return bounded_records(_normalize_record(row) for row in csv.DictReader(handle))
    if suffix in {".xlsx", ".xls"}:
        if suffix == ".xls":
            raise IngestionReadError("Legacy .xls files are not supported; use .xlsx")
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value) if value is not None else "" for value in next(rows)]
        except StopIteration:
            return [], 0
        if any(not header for header in headers) or len(set(headers)) != len(headers):
            raise IngestionReadError("Excel header names must be non-empty and unique")
        try:
            return bounded_records(dict(zip(headers, row)) for row in rows)
        finally:
            workbook.close()
    if suffix == ".json":
        with source.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        if not isinstance(payload, list):
            raise IngestionReadError("JSON file root must be an array of objects")
        return bounded_records(_normalize_record(row) for row in payload)
    raise IngestionReadError(f"Unsupported file format: {suffix or '(none)'}")


def _extract_path(payload: Any, path: str | None) -> Any:
    current = payload
    if not path:
        return current
    for segment in path.split("."):
        if not isinstance(current, Mapping) or segment not in current:
            raise IngestionReadError(f"records_path segment not found: {segment}")
        current = current[segment]
    return current


def read_api_records(
    config_data: Mapping[str, Any],
    credential: CredentialWrite | None = None,
) -> tuple[List[Dict[str, Any]], int]:
    config = ApiSourceConfig.model_validate(config_data)
    headers = dict(config.headers)
    if credential and credential.credential_type.value == "API_KEY":
        headers[str(credential.header_name)] = str(credential.secret)
    elif credential and credential.credential_type.value == "BEARER":
        headers["Authorization"] = f"Bearer {credential.secret}"

    def pages() -> Iterable[Dict[str, Any]]:
        page = config.pagination.start_page
        page_count = 0
        while True:
            query = dict(config.query)
            if config.pagination.mode == "PAGE":
                query[config.pagination.page_param] = str(page)
                query[config.pagination.page_size_param] = str(config.pagination.page_size)
            response = requests.get(
                str(config.url),
                params=query,
                headers=headers,
                timeout=config.timeout_seconds,
            )
            if response.status_code >= 400:
                raise IngestionReadError(
                    f"Source API returned HTTP {response.status_code}"
                )
            try:
                values = _extract_path(response.json(), config.records_path)
            except ValueError as exc:
                raise IngestionReadError("Source API did not return valid JSON") from exc
            if not isinstance(values, list):
                raise IngestionReadError("Configured records_path must resolve to an array")
            for value in values:
                yield _normalize_record(value)
            page_count += 1
            if config.pagination.mode == "NONE" or not values:
                break
            if len(values) < config.pagination.page_size:
                break
            if page_count >= config.pagination.max_pages:
                raise IngestionReadError("Source API exceeded configured max_pages")
            page += 1

    return bounded_records(pages())
